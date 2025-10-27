#--------------------------CRIANDO ARQUIVO EXCELL---------------------------------
'''from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Estoque"               #aba inicial (semopre tem uma "sheet" por padrao, entao modifica o nome)
ws2 = wb.create_sheet("Casatro")   #adicionando abas a mais
ws3 = wb.create_sheet("Lojas", 0)

ws3.append(["Produto", "Quantidade", "Preço"])
ws3.append(["Açucar", 5, 10])
ws3.append(["café", 15, 20])

ws["D1"] = "Total"                 #adicionar em um quadrado especifico
wb.save("estoque.xlsx")            #salvar'''




#--------------------------ABRINDO ARQUIVO JA SALVO-----------------------------------
'''from openpyxl import load_workbook
wb = load_workbook(r"estoque.xlsx")
ws = wb["Lojas"]

print(ws["B2"].value)'''



#-----MANIPULAÇÃO DAS LINHAS-----
'''for linha in ws.iter_rows(min_row=2, values_only=True):     #min = onde vai comer o loop
    print(linha)
    produto, qtd, preco = linha
    print(f"o produto {produto} possui {qtd} unidades, no valor de {preco} reais")

wb.save("estoque.xlsx")'''



#---------------------------DESAFIO-----------------------------
#-------------------------
import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Entrada"
ws2 = wb.create_sheet("Saídas")

ws.append(["Produto", "Preço"])
ws2.append(["Produto", "Preço"])

wb.save("estoquedesafio.xlsx")

while True:
    print(f"Bem vindo a lista de chamados, selecione o que deseja:")
    try:
        n = int(input(f" 1- Adicionar entradas\n 2- Adicionar Saídas\n 3- Vizualizar entradas\n 4- Vizualizar Saidas\n 5- Total de entradas e saídas\n 0- Encerrar"))

        if n == 1:
            print("-----------------------------------------------------------------------")
            Produto = input("Informe qual seria o produto:")
            preco = int(input("Informe o preço do produto:"))
            ws.append([Produto, preco])
            wb.save("estoquedesafio.xlsx")
            continue

        if n == 2:
            print("-----------------------------------------------------------------------")
            Produto = input("Informe qual seria o produto:")
            preco = int(input("Informe o preço do produto:"))
            ws2.append([Produto, preco])
            wb.save("estoquedesafio.xlsx")
            continue

        if n == 3:
            print("-----------------------------------------------------------------------")
            for linha in ws.iter_rows(min_row=2, values_only=True):
                print(linha)
                produto, preço = linha
                print(f"{produto} no valor de {preço} reais")
                continue

        if n == 4:
            print("-----------------------------------------------------------------------")
            for linha in ws2.iter_rows(min_row=2, values_only=True):
                print(linha)
                produto, preço = linha
                print(f"{produto} no valor de {preço} reais")
                continue

        if n == 5:
            print("-----------------------------------------------------------------------")
            #somar tudo

        if n == 0:
            print(f"encerrando...")
            break

        else:
            print("Você deve digitar um dos numeros da lista\n")
            continue
    except ValueError:
        print("voce deve digitar um número..")